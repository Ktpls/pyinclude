from .util_solid import *

"""
numpy
"""

import numpy as np
import hashlib

np.seterr(all="raise")
import scipy.interpolate as interpolate

EPS = 1e-10
LOGEPS = -10


def forceviewmaxmin(m):
    ma = m.max()
    mi = m.min()
    print("ma", ma)
    print("mi", mi)
    pass


def arrayshift(a, n, fill=np.nan):
    # n positive as right
    if n == 0:
        return a
    elif n > 0:
        # a[:-n] will not work as intended on n==0
        return np.concatenate((np.full(n, fill), a[:-n]))
    else:
        return np.concatenate((a[-n:], np.full(-n, fill)))


def integral(dx, x0, keepXM1=False):
    # keepXM1 to keep the last element in x, or deprecate it, cuz its nan if generated from derivative
    x = np.array(list(itertools.accumulate(dx, lambda t, e: t + e)))
    x = np.concatenate([[x0], x if keepXM1 else x[:-1]])
    return x


def derivative(x):
    return arrayshift(x, -1) - x


def summonCard(inteprob, generator=None):
    # summon from card pool
    # impl using np
    # norm
    prob = np.array(inteprob, dtype=np.float32)
    prob /= prob.sum()
    if generator is None:
        return np.random.choice(np.arange(len(prob)), p=prob)
    else:
        return generator.choice(np.arange(len(prob)), p=prob)


class ZFunc:
    """
    x1x2 at any order
    """

    def __init__(self, x1, y1, x2, y2) -> None:
        assert np.fabs(x1 - x2) > EPS
        if x1 < x2:
            ptleft = np.array([x1, y1])
            ptright = np.array([x2, y2])
        else:
            ptright = np.array([x1, y1])
            ptleft = np.array([x2, y2])
        self.yminmax = np.array([min(y1, y2), max(y1, y2)])
        self.slope = (ptright[1] - ptleft[1]) / (ptright[0] - ptleft[0])
        self.bias = ptleft[1] - self.slope * ptleft[0]

    def __CallOnNDArray(self, x: np.ndarray):
        y = self.slope * x + self.bias
        y = np.clip(y, self.yminmax[0], self.yminmax[1])
        return y

    def __CallOnNum(self, x):
        y = self.slope * x + self.bias
        if y < self.yminmax[0]:
            y = self.yminmax[0]
        elif y > self.yminmax[1]:
            y = self.yminmax[1]
        return y

    def __call__(self, x):
        if type(x) is np.ndarray:
            return self.__CallOnNDArray(x)
        else:
            return self.__CallOnNum(x)


def randomString(charset, length):
    return "".join(
        [
            charset[i]
            for i in np.random.choice(range(len(charset)), length, replace=True)
        ]
    )


def ReLU(x):
    return np.maximum(0, x)


def SafeExp(x):
    x[x < LOGEPS] = LOGEPS
    y = np.exp(x)
    y[x < LOGEPS] = 0
    return y


def SafeLog(x):
    x[x < EPS] = EPS
    y = np.log(x)
    return y


def NormalizeIterableOrSingleArgToNdarray(x):
    if type(x) is np.ndarray:
        return x
    return np.array(NormalizeIterableOrSingleArgToIterable(x))


@dataclasses.dataclass
class BayesEstimator:
    xspace: np.ndarray
    distributionModel: typing.Callable  # to calc P(measuredVal=B|val=A)
    logPBASum: np.ndarray = dataclasses.field(init=False, default=None)

    logSumLowerLimit = -500

    def __post_init__(self):
        self.logPBASum = np.zeros_like(self.xspace)

    def update(self, measuredValue: float | list[float] | np.ndarray):
        measuredValue = NormalizeIterableOrSingleArgToNdarray(measuredValue)
        measuredValue = np.array(measuredValue)
        PBA = self.distributionModel(
            self.xspace.reshape((-1, 1)), measuredValue.reshape((1, -1))
        )
        self.logPBASum += np.sum(SafeLog(PBA), axis=1)
        self.logPBASum -= np.max(self.logPBASum)
        self.logPBASum[self.logPBASum < BayesEstimator.logSumLowerLimit] = (
            BayesEstimator.logSumLowerLimit
        )

    def getPossibility(self):
        possibility = SafeExp(self.logPBASum)
        possibility /= np.sum(possibility)
        return possibility


"""
xls
"""

import openpyxl as opx


def save_list_to_xls(
    data_list: list[list | tuple | typing.Any], filename, sheetname=None
):
    # Create a new workbook
    wb = opx.Workbook()

    # Select the active worksheet
    if sheetname is None:
        ws = wb.active
    else:
        ws = wb.create_sheet(sheetname)

    # Iterate over the list and write each item to a new row
    for row, rowcontent in enumerate(data_list):
        rowcontent = NormalizeIterableOrSingleArgToIterable(rowcontent)
        for col, item in enumerate(rowcontent):
            ws.cell(row=row + 1, column=col + 1, value=item)

    # Save the workbook to the specified filename
    wb.save(filename)


def Xls2ListList(path=None, sheetname=None, killNones=True):
    if path is None:
        path = r"eles.in.xlsx"
    xls = opx.load_workbook(path)
    if sheetname is None:
        sheet = xls.active
    else:
        sheet = xls[sheetname]

    ret = [[ele.value for ele in ln] for ln in (sheet.rows)]
    if killNones:
        ret = [l for l in ret if any([e is not None for e in l])]
    return ret


def NpGeneratorFromStrSeed(s: str):
    seed = s.encode("utf-8")
    seed = hashlib.md5(seed).digest()
    seed = int.from_bytes(seed[:8])
    return np.random.Generator(np.random.PCG64(seed))
