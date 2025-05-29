from .util_solid import NormalizeIterableOrSingleArgToIterable
import typing
import openpyxl as opx

"""
xls
"""


def save_list_to_xls(
    data_list: list[list | tuple | typing.Any], filename, sheetname=None
):
    # Create a new workbook
    wb = opx.Workbook()

    # Select the active worksheet
    if sheetname is None:
        ws = wb.active
    else:
        ws = wb.create_sheet(sheetname)

    # Iterate over the list and write each item to a new row
    for row, rowcontent in enumerate(data_list):
        rowcontent = NormalizeIterableOrSingleArgToIterable(rowcontent)
        for col, item in enumerate(rowcontent):
            ws.cell(row=row + 1, column=col + 1, value=item)

    # Save the workbook to the specified filename
    wb.save(filename)


def Xls2ListList(path=None, sheetname=None, killNones=True):
    if path is None:
        path = r"eles.in.xlsx"
    xls = opx.load_workbook(path)
    if sheetname is None:
        sheet = xls.active
    else:
        sheet = xls[sheetname]

    ret = [[ele.value for ele in ln] for ln in (sheet.rows)]
    if killNones:
        ret = [l for l in ret if any([e is not None for e in l])]
    return ret
